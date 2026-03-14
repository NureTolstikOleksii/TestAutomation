from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import re

wb = load_workbook("excel-files/data.xlsx")
sheet = wb.active

green_font = Font(color="008000")
blue_font = Font(color="0000FF")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

for row in sheet.iter_rows():
    # 1: Перевіряємо, чи є в усьому рядку хоча б одна буква
    row_contains_letters = False
    for cell in row:
        if cell.value is not None and re.search(r"[a-zA-Zа-яА-ЯіІїЇєЄґҐ]", str(cell.value)):
            row_contains_letters = True
            break

    # 2: Застосовуємо форматування до кожної комірки в рядку
    for cell in row:
        if cell.value is None:
            continue

        val_str = str(cell.value)
        has_letters = bool(re.search(r"[a-zA-Zа-яА-ЯіІїЇєЄґҐ]", val_str))
        has_digits = bool(re.search(r"\d", val_str))

        if row_contains_letters:
            cell.font = blue_font

        if has_digits and not has_letters:
            cell.font = green_font

        if has_letters and has_digits:
            cell.fill = yellow_fill

wb.save("excel-files/result.xlsx")
print("Done! Result in file result.xlsx")